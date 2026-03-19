from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROWS = [
    {
        "策略名称": "动量策略",
        "strategy_type": "momentum",
        "模板风格": "JoinQuant",
        "关键参数": "lookback_days, stock_count",
        "推荐提示词": "最近60天涨幅最高的20只股票",
    },
    {
        "策略名称": "均线突破",
        "strategy_type": "ma_breakout",
        "模板风格": "JoinQuant",
        "关键参数": "ma_period, threshold, stock_code",
        "推荐提示词": "5日均线突破1%买入平安银行",
    },
    {
        "策略名称": "KDJ择时",
        "strategy_type": "kdj_timing",
        "模板风格": "JoinQuant",
        "关键参数": "stock_code, k_period, buy_threshold, sell_threshold",
        "推荐提示词": "KDJ低于20买入，高于80卖出",
    },
    {
        "策略名称": "CTA趋势突破",
        "strategy_type": "cta",
        "模板风格": "JoinQuant",
        "关键参数": "symbol_code, breakout_days, exit_days",
        "推荐提示词": "做一个CTA策略，交易沪深300ETF，突破过去20天最高价买入，跌破过去10天最低价卖出",
    },
    {
        "策略名称": "趋势跟踪",
        "strategy_type": "trend_following",
        "模板风格": "JoinQuant",
        "关键参数": "stock_code, fast_ma, slow_ma",
        "推荐提示词": "做一个趋势跟踪策略，交易平安银行，5日均线上穿20日均线买入，下穿卖出",
    },
    {
        "策略名称": "布林带突破",
        "strategy_type": "bollinger_breakout",
        "模板风格": "JoinQuant",
        "关键参数": "stock_code, period, num_std",
        "推荐提示词": "做一个布林带突破策略，交易平安银行，20日布林带，上轨突破买入，下轨跌破卖出，标准差2倍",
    },
    {
        "策略名称": "随机森林选股",
        "strategy_type": "random_forest",
        "模板风格": "BigQuant",
        "关键参数": "train_start_date, train_end_date, predict_start_date, predict_end_date, stock_count, n_estimators",
        "推荐提示词": "做一个随机森林选股策略，训练区间是2020年，预测区间是2021年，买入预测分数最高的10只股票，随机森林树数量10",
    },
    {
        "策略名称": "羊驼随机轮动",
        "strategy_type": "alpaca_rotation",
        "模板风格": "BigQuant",
        "关键参数": "total_stock_nums, sell_stock_nums, rebalance_days, start_date, end_date, random_seed",
        "推荐提示词": "做一个羊驼策略，随机持有30只股票，每22个交易日调仓一次，卖出收益最差的6只股票，再随机买入6只，回测区间2009年到2021年",
    },
    {
        "策略名称": "布兰德价值投资",
        "strategy_type": "brandes_value",
        "模板风格": "BigQuant",
        "关键参数": "hold_count, start_date, end_date, rebalance_period_days",
        "推荐提示词": "做一个布兰德价值投资策略，持股30只，每月第1个交易日调仓，回测区间2016年4月到2024年9月",
    },
]


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    wrap = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_workbook() -> Workbook:
    wb = Workbook()

    strategy_ws = wb.active
    strategy_ws.title = "Strategies"
    headers = list(ROWS[0].keys())
    strategy_ws.append(headers)
    for row in ROWS:
        strategy_ws.append([row[h] for h in headers])

    style_sheet(strategy_ws)
    strategy_ws.column_dimensions["A"].width = 18
    strategy_ws.column_dimensions["B"].width = 24
    strategy_ws.column_dimensions["C"].width = 14
    strategy_ws.column_dimensions["D"].width = 60
    strategy_ws.column_dimensions["E"].width = 90
    for idx in range(2, strategy_ws.max_row + 1):
        strategy_ws.row_dimensions[idx].height = 42

    prompt_ws = wb.create_sheet("Prompt List")
    prompt_ws.append(["序号", "策略名称", "推荐提示词"])
    for idx, row in enumerate(ROWS, start=1):
        prompt_ws.append([idx, row["策略名称"], row["推荐提示词"]])

    style_sheet(prompt_ws)
    prompt_ws.column_dimensions["A"].width = 8
    prompt_ws.column_dimensions["B"].width = 18
    prompt_ws.column_dimensions["C"].width = 100
    for idx in range(2, prompt_ws.max_row + 1):
        prompt_ws.row_dimensions[idx].height = 36

    return wb


def main() -> None:
    output = Path(__file__).resolve().parent / "STRATEGY_PROMPTS.xlsx"
    workbook = build_workbook()
    workbook.save(output)
    print(output)


if __name__ == "__main__":
    main()
